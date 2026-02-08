"""
Import service
Handles bulk import of invitees from Excel/CSV files
"""
import pandas as pd
import re
from flask import request
from app import db
from app.models.invitee import Invitee
from app.models.event_invitee import EventInvitee
from app.models.user import User
from app.models.audit_log import AuditLog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

class ImportService:
    """Service for bulk importing invitees"""
    
    @staticmethod
    def validate_email(email):
        """Validate email format"""
        pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return re.match(pattern, str(email)) is not None
    
    @staticmethod
    def validate_phone(phone):
        """Validate phone format: must start with 201 and be 12 digits"""
        clean_phone = str(phone).replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        pattern = r'^201\d{9}$'
        return re.match(pattern, clean_phone) is not None
    
    @staticmethod
    def import_contacts_from_file(filepath, user_id):
        """
        Import contacts from Excel or CSV file.
        Extracts: name, email, phone, inviter, category, allowed_guests, etc.
        Mandatory: name, email, phone, inviter. Others optional.
        Validates phone: must start with 201 and be 12 digits. Skips invalid entries.
        Returns dict with import results.
        """
        # Read file based on extension
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

        # Normalize column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        # Debug: Print actual column names found
        print(f"DEBUG: Original columns: {list(df.columns)}")
        print(f"DEBUG: Normalized columns: {list(df.columns)}")

        # Validate required columns
        required_columns = ['name', 'email', 'phone', 'inviter']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        # Try alternative column name patterns if some are missing
        if missing_columns:
            print(f"DEBUG: Missing columns: {missing_columns}")
            
            # Check if columns exist with different spacing/casing
            for col in missing_columns:
                if col == 'inviter':
                    # Check for 'inviter_name' or other variations
                    if 'inviter_name' in df.columns:
                        df.rename(columns={'inviter_name': 'inviter'}, inplace=True)
                        missing_columns.remove(col)
                        print(f"DEBUG: Renamed 'inviter_name' to 'inviter'")
                    elif 'inviter' not in df.columns and 'inviter' in [c.lower() for c in df.columns]:
                        # Case insensitive match
                        for c in df.columns:
                            if c.lower() == 'inviter':
                                df.rename(columns={c: 'inviter'}, inplace=True)
                                missing_columns.remove(col)
                                print(f"DEBUG: Renamed '{c}' to 'inviter'")
                                break
            
            # Re-check after renaming
            missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise ValueError(
                f"Missing required columns: {', '.join(missing_columns)}. "
                "Please make sure the first row contains properly named column headers: Name, Email, Phone, Inviter. "
                "Your file appears to have data in the first row instead of headers."
            )

        # Get user info for audit
        user = User.query.get(user_id)
        if not user:
            raise ValueError("User not found")

        # Process rows
        total_rows = len(df)
        successful = 0
        skipped = 0
        failed = 0
        errors = []

        from app.models.inviter import Inviter
        for index, row in df.iterrows():
            try:
                # Extract fields
                name = str(row['name']).strip() if pd.notna(row['name']) else None
                email = str(row['email']).strip().lower() if pd.notna(row['email']) else None
                # Clean phone: remove whitespace and common separators
                raw_phone = str(row['phone']).strip() if pd.notna(row['phone']) else None
                phone = raw_phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '') if raw_phone else None
                inviter_name = str(row['inviter']).strip() if pd.notna(row['inviter']) else None
                category_name = str(row['category']).strip() if 'category' in row and pd.notna(row['category']) else None
                category_id = None
                from app.models.category import Category
                if category_name:
                    category_obj = Category.query.filter_by(name=category_name).first()
                    if category_obj:
                        category_id = category_obj.id
                    else:
                        # Category not found - leave as null (no category)
                        category_id = None
                        errors.append(f"Row {index + 2}: Category '{category_name}' not found, imported without category.")
                # Empty category - leave as null (no category assigned)
                allowed_guests = int(row['allowed_guests']) if 'allowed_guests' in row and pd.notna(row['allowed_guests']) else None
                position = str(row['position']).strip() if 'position' in row and pd.notna(row['position']) else None
                company = str(row['company']).strip() if 'company' in row and pd.notna(row['company']) else None

                # Validate mandatory fields
                if not name or not email or not phone or not inviter_name:
                    skipped += 1
                    errors.append(f"Row {index + 2}: Missing required field (name, email, phone, or inviter)")
                    continue

                # Validate phone format
                if not ImportService.validate_phone(phone):
                    skipped += 1
                    errors.append(f"Row {index + 2}: Invalid phone format (must start with 201 and be 12 digits)")
                    continue

                # Email validation commented out - no need to validate email
                # if not ImportService.validate_email(email):
                #     skipped += 1
                #     errors.append(f"Row {index + 2}: Invalid email format")
                #     continue

                # Find inviter by name in user's group, create if not found
                inviter_obj = Inviter.query.filter_by(
                    name=inviter_name,
                    inviter_group_id=user.inviter_group_id
                ).first()
                if not inviter_obj:
                    inviter_obj = Inviter(
                        name=inviter_name,
                        inviter_group_id=user.inviter_group_id,
                        is_active=True
                    )
                    db.session.add(inviter_obj)
                    db.session.flush()  # Get ID

                inviter_id = inviter_obj.id

                # Check if contact already exists by phone within this inviter group
                existing_invitee = Invitee.find_by_phone_in_group(phone, user.inviter_group_id)

                if existing_invitee:
                    # Contact already exists in this group - update with new data
                    updated_fields = []
                    
                    # Update name if different
                    if name and existing_invitee.name != name:
                        existing_invitee.name = name
                        updated_fields.append('name')
                    
                    # Update email if different
                    if email and existing_invitee.email != email:
                        existing_invitee.email = email
                        updated_fields.append('email')
                    
                    # Update position if provided
                    if position and existing_invitee.position != position:
                        existing_invitee.position = position
                        updated_fields.append('position')
                    
                    # Update company if provided
                    if company and existing_invitee.company != company:
                        existing_invitee.company = company
                        updated_fields.append('company')
                    
                    # Update category if provided
                    if category_id and existing_invitee.category_id != category_id:
                        existing_invitee.category_id = category_id
                        updated_fields.append('category')
                    
                    # Update plus_one (allowed guests) if provided
                    if allowed_guests is not None and existing_invitee.plus_one != allowed_guests:
                        existing_invitee.plus_one = allowed_guests
                        updated_fields.append('plus_one')
                    
                    # Update inviter if different
                    if inviter_id and existing_invitee.inviter_id != inviter_id:
                        existing_invitee.inviter_id = inviter_id
                        updated_fields.append('inviter')
                    
                    if updated_fields:
                        successful += 1
                        errors.append(f"Row {index + 2}: Updated existing contact '{phone}' ({', '.join(updated_fields)})")
                    else:
                        skipped += 1
                        errors.append(f"Row {index + 2}: Contact with phone '{phone}' already exists (no changes)")
                    continue

                # Create new contact
                invitee = Invitee(
                    name=name,
                    email=email,
                    phone=phone,
                    position=position,
                    company=company,
                    category_id=category_id,
                    plus_one=allowed_guests,  # Use plus_one field
                    inviter_group_id=user.inviter_group_id,  # Assign to user's group
                    inviter_id=inviter_id
                )
                db.session.add(invitee)
                successful += 1

            except Exception as e:
                failed += 1
                errors.append(f"Row {index + 2}: {str(e)}")

        # Commit all successful changes
        try:
            db.session.commit()
            # Log import
            AuditLog.log(
                user_id=user_id,
                action='bulk_import_contacts',
                table_name='invitees',
                new_value=f'Imported {successful} contacts, {skipped} skipped, {failed} failed',
                ip_address=request.remote_addr if request else None
            )
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Database error during import: {str(e)}")

        return {
            'total_rows': total_rows,
            'successful': successful,
            'skipped': skipped,
            'failed': failed,
            'errors': errors
        }
    
    @staticmethod
    def admin_import_contacts_from_file(filepath, user_id):
        """
        Admin bulk import: application-wide.
        Same logic as group-scoped import but each row specifies its Inviter_Group.
        Groups must already exist. Inviters auto-created within the specified group.
        """
        if filepath.endswith('.csv'):
            df = pd.read_csv(filepath)
        else:
            df = pd.read_excel(filepath)

        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')

        # Validate required columns (inviter_group is the new required column)
        required_columns = ['inviter_group', 'name', 'email', 'phone', 'inviter']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            # Try alternative names
            for col in list(missing_columns):
                if col == 'inviter_group':
                    for alt in ['inviter_group_name', 'group', 'group_name']:
                        if alt in df.columns:
                            df.rename(columns={alt: 'inviter_group'}, inplace=True)
                            missing_columns.remove(col)
                            break
                elif col == 'inviter':
                    if 'inviter_name' in df.columns:
                        df.rename(columns={'inviter_name': 'inviter'}, inplace=True)
                        missing_columns.remove(col)

            missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValueError(
                f"Missing required columns: {', '.join(missing_columns)}. "
                "Required headers: Inviter_Group, Name, Email, Phone, Inviter."
            )

        user = User.query.get(user_id)
        if not user:
            raise ValueError("User not found")

        from app.models.inviter import Inviter
        from app.models.inviter_group import InviterGroup
        from app.models.category import Category

        total_rows = len(df)
        successful = 0
        skipped = 0
        failed = 0
        errors = []

        # Cache lookups to avoid repeated DB queries
        group_cache = {}  # name -> InviterGroup or None
        inviter_cache = {}  # (group_id, inviter_name) -> Inviter

        for index, row in df.iterrows():
            try:
                # Extract inviter group
                group_name = str(row['inviter_group']).strip() if pd.notna(row['inviter_group']) else None
                name = str(row['name']).strip() if pd.notna(row['name']) else None
                email = str(row['email']).strip().lower() if pd.notna(row['email']) else None
                raw_phone = str(row['phone']).strip() if pd.notna(row['phone']) else None
                phone = raw_phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '') if raw_phone else None
                inviter_name = str(row['inviter']).strip() if pd.notna(row['inviter']) else None
                category_name = str(row['category']).strip() if 'category' in row and pd.notna(row['category']) else None
                category_id = None
                if category_name:
                    category_obj = Category.query.filter_by(name=category_name).first()
                    if category_obj:
                        category_id = category_obj.id
                    else:
                        errors.append(f"Row {index + 2}: Category '{category_name}' not found, imported without category.")
                allowed_guests = int(row['allowed_guests']) if 'allowed_guests' in row and pd.notna(row['allowed_guests']) else None
                position = str(row['position']).strip() if 'position' in row and pd.notna(row['position']) else None
                company = str(row['company']).strip() if 'company' in row and pd.notna(row['company']) else None

                # Validate mandatory fields
                if not group_name or not name or not email or not phone or not inviter_name:
                    skipped += 1
                    errors.append(f"Row {index + 2}: Missing required field (inviter_group, name, email, phone, or inviter)")
                    continue

                # Validate phone format
                if not ImportService.validate_phone(phone):
                    skipped += 1
                    errors.append(f"Row {index + 2}: Invalid phone format (must start with 201 and be 12 digits)")
                    continue

                # Look up inviter group by name (must already exist)
                if group_name not in group_cache:
                    group_cache[group_name] = InviterGroup.get_by_name(group_name)
                group_obj = group_cache[group_name]

                if not group_obj:
                    skipped += 1
                    errors.append(f"Row {index + 2}: Inviter group '{group_name}' not found. Please create it first.")
                    continue

                group_id = group_obj.id

                # Find or create inviter within the group
                cache_key = (group_id, inviter_name)
                if cache_key not in inviter_cache:
                    inviter_obj = Inviter.query.filter_by(
                        name=inviter_name,
                        inviter_group_id=group_id
                    ).first()
                    if not inviter_obj:
                        inviter_obj = Inviter(
                            name=inviter_name,
                            inviter_group_id=group_id,
                            is_active=True
                        )
                        db.session.add(inviter_obj)
                        db.session.flush()
                    inviter_cache[cache_key] = inviter_obj
                inviter_obj = inviter_cache[cache_key]
                inviter_id = inviter_obj.id

                # Check if contact already exists by phone within this group
                existing_invitee = Invitee.find_by_phone_in_group(phone, group_id)

                if existing_invitee:
                    updated_fields = []
                    if name and existing_invitee.name != name:
                        existing_invitee.name = name
                        updated_fields.append('name')
                    if email and existing_invitee.email != email:
                        existing_invitee.email = email
                        updated_fields.append('email')
                    if position and existing_invitee.position != position:
                        existing_invitee.position = position
                        updated_fields.append('position')
                    if company and existing_invitee.company != company:
                        existing_invitee.company = company
                        updated_fields.append('company')
                    if category_id and existing_invitee.category_id != category_id:
                        existing_invitee.category_id = category_id
                        updated_fields.append('category')
                    if allowed_guests is not None and existing_invitee.plus_one != allowed_guests:
                        existing_invitee.plus_one = allowed_guests
                        updated_fields.append('plus_one')
                    if inviter_id and existing_invitee.inviter_id != inviter_id:
                        existing_invitee.inviter_id = inviter_id
                        updated_fields.append('inviter')

                    if updated_fields:
                        successful += 1
                        errors.append(f"Row {index + 2}: Updated existing contact '{phone}' ({', '.join(updated_fields)})")
                    else:
                        skipped += 1
                        errors.append(f"Row {index + 2}: Contact with phone '{phone}' already exists (no changes)")
                    continue

                # Create new contact
                invitee = Invitee(
                    name=name,
                    email=email,
                    phone=phone,
                    position=position,
                    company=company,
                    category_id=category_id,
                    plus_one=allowed_guests,
                    inviter_group_id=group_id,
                    inviter_id=inviter_id
                )
                db.session.add(invitee)
                successful += 1

            except Exception as e:
                failed += 1
                errors.append(f"Row {index + 2}: {str(e)}")

        try:
            db.session.commit()
            AuditLog.log(
                user_id=user_id,
                action='bulk_import_contacts',
                table_name='invitees',
                new_value=f'Admin import: {successful} contacts, {skipped} skipped, {failed} failed',
                ip_address=request.remote_addr if request else None
            )
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            raise Exception(f"Database error during import: {str(e)}")

        return {
            'total_rows': total_rows,
            'successful': successful,
            'skipped': skipped,
            'failed': failed,
            'errors': errors
        }

    @staticmethod
    def generate_template():
        """
        Generate Excel template for contact import.
        Returns the file path of the generated template.
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Instructions sheet
        instructions_ws = wb.create_sheet("Instructions", 0)

        # Title
        instructions_ws['A1'] = "BULK IMPORT INSTRUCTIONS"
        instructions_ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        instructions_ws['A1'].fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        instructions_ws['A1'].alignment = Alignment(horizontal="center")
        instructions_ws.merge_cells('A1:J1')

        # Instructions content
        instructions = [
            [],
            ["REQUIRED COLUMNS", "", "", "", "", "", "", "", "", ""],
            ["Column Name", "Required?", "Description", "Example", "", "", "", "", "", ""],
            ["Name", "YES", "Full name of the invitee", "John Smith", "", "", "", "", "", ""],
            ["Email", "YES", "Valid email address", "john@example.com", "", "", "", "", "", ""],
            ["Phone", "YES", "Phone number (must start with 201 and be 12 digits)", "201012345678", "", "", "", "", "", ""],
            ["Inviter", "YES", "Name of inviter (person or group)", "Ali Hassan", "", "", "", "", "", ""],
            [],
            ["OPTIONAL COLUMNS", "", "", "", "", "", "", "", "", ""],
            ["Column Name", "Required?", "Description", "Example", "", "", "", "", "", ""],
            ["Category", "NO", "Invitee category (White/Gold)", "White", "", "", "", "", "", ""],
            ["Allowed_Guests", "NO", "Guests allowed for this invitee", "2", "", "", "", "", "", ""],
            ["Position", "NO", "Job title or position", "Manager", "", "", "", "", "", ""],
            ["Company", "NO", "Company name", "ABC Corporation", "", "", "", "", "", ""],
            [],
            ["IMPORTANT NOTES", "", "", "", "", "", "", "", "", ""],
            ["1. Do not change the column headers in the template", "", "", "", "", "", "", "", "", ""],
            ["2. All rows with missing Name, Email, Phone, or Inviter will be skipped", "", "", "", "", "", "", "", "", ""],
            ["3. Phone must start with 201 and be 12 digits (e.g., 201012345678)", "", "", "", "", "", "", "", "", ""],
            ["4. If a contact with the same phone already exists, the row will be skipped or updated", "", "", "", "", "", "", "", "", ""],
            ["5. After importing contacts, go to the Events tab to submit them to events", "", "", "", "", "", "", "", "", ""],
            [],
            ["Go to the 'Template' sheet to start importing data →", "", "", "", "", "", "", "", "", ""],
        ]

        for i, row_data in enumerate(instructions, start=1):
            for j, cell_value in enumerate(row_data, start=1):
                cell = instructions_ws.cell(row=i, column=j, value=cell_value)

                # Style headers
                if i == 2 or i == 9:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                elif i == 3 or i == 10:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                elif i == 14:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                elif i == 20:
                    cell.font = Font(bold=True, size=12, color="3B82F6")

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            instructions_ws.column_dimensions[col].width = 20

        # Create Template sheet
        template_ws = wb.create_sheet("Template", 1)

        # Headers (contacts only - no event-specific fields)
        headers = ["Name", "Email", "Phone", "Inviter", "Category", "Allowed_Guests", "Position", "Company"]
        for col, header in enumerate(headers, start=1):
            cell = template_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Add sample data
        sample_data = [
            ["Ahmed Hassan", "ahmed.hassan@example.com", "201012345678", "Ali Hassan", "White", 2, "CEO", "Tech Solutions"],
            ["Sarah Mohamed", "sarah.mohamed@example.com", "201123456789", "Fatma Ali", "Gold", 1, "Marketing Director", "Creative Agency"],
            ["Mohamed Ali", "mohamed.ali@example.com", "201234567890", "Omar Said", "White", 0, "Sales Manager", "Sales Corp"],
        ]

        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                template_ws.cell(row=row_idx, column=col_idx, value=value)

        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            template_ws.column_dimensions[col].width = 20

        # Save template - use absolute path based on app root
        import tempfile
        template_path = os.path.join(tempfile.gettempdir(), 'contacts_import_template.xlsx')
        wb.save(template_path)

        return template_path

    @staticmethod
    def generate_admin_template():
        """
        Generate Excel template for admin-wide contact import.
        Includes Inviter_Group as a required column.
        """
        wb = Workbook()
        wb.remove(wb.active)

        # Instructions sheet
        instructions_ws = wb.create_sheet("Instructions", 0)
        instructions_ws['A1'] = "ADMIN BULK IMPORT INSTRUCTIONS"
        instructions_ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        instructions_ws['A1'].fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        instructions_ws['A1'].alignment = Alignment(horizontal="center")
        instructions_ws.merge_cells('A1:J1')

        instructions = [
            [],
            ["REQUIRED COLUMNS", "", "", "", "", "", "", "", "", ""],
            ["Column Name", "Required?", "Description", "Example", "", "", "", "", "", ""],
            ["Inviter_Group", "YES", "Name of the inviter group (must already exist)", "Marketing Team", "", "", "", "", "", ""],
            ["Name", "YES", "Full name of the invitee", "John Smith", "", "", "", "", "", ""],
            ["Email", "YES", "Valid email address", "john@example.com", "", "", "", "", "", ""],
            ["Phone", "YES", "Phone number (must start with 201 and be 12 digits)", "201012345678", "", "", "", "", "", ""],
            ["Inviter", "YES", "Name of inviter (auto-created if new in group)", "Ali Hassan", "", "", "", "", "", ""],
            [],
            ["OPTIONAL COLUMNS", "", "", "", "", "", "", "", "", ""],
            ["Column Name", "Required?", "Description", "Example", "", "", "", "", "", ""],
            ["Category", "NO", "Invitee category (White/Gold)", "White", "", "", "", "", "", ""],
            ["Allowed_Guests", "NO", "Guests allowed for this invitee", "2", "", "", "", "", "", ""],
            ["Position", "NO", "Job title or position", "Manager", "", "", "", "", "", ""],
            ["Company", "NO", "Company name", "ABC Corporation", "", "", "", "", "", ""],
            [],
            ["IMPORTANT NOTES", "", "", "", "", "", "", "", "", ""],
            ["1. Inviter groups must exist before import — create them in the Groups page", "", "", "", "", "", "", "", "", ""],
            ["2. Inviters are automatically created within the specified group if they don't exist", "", "", "", "", "", "", "", "", ""],
            ["3. Phone must start with 201 and be 12 digits (e.g., 201012345678)", "", "", "", "", "", "", "", "", ""],
            ["4. If a contact with the same phone already exists in the group, it will be updated", "", "", "", "", "", "", "", "", ""],
            ["5. You can mix contacts from different groups in one file", "", "", "", "", "", "", "", "", ""],
            [],
            ["Go to the 'Template' sheet to start importing data →", "", "", "", "", "", "", "", "", ""],
        ]

        for i, row_data in enumerate(instructions, start=1):
            for j, cell_value in enumerate(row_data, start=1):
                cell = instructions_ws.cell(row=i, column=j, value=cell_value)
                if i == 2 or i == 10:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                elif i == 3 or i == 11:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                elif i == 16:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                elif i == 22:
                    cell.font = Font(bold=True, size=12, color="7C3AED")

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            instructions_ws.column_dimensions[col].width = 22

        # Template sheet
        template_ws = wb.create_sheet("Template", 1)
        headers = ["Inviter_Group", "Name", "Email", "Phone", "Inviter", "Category", "Allowed_Guests", "Position", "Company"]
        for col, header in enumerate(headers, start=1):
            cell = template_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        sample_data = [
            ["Marketing Team", "Ahmed Hassan", "ahmed@example.com", "201012345678", "Ali Hassan", "White", 2, "CEO", "Tech Solutions"],
            ["Marketing Team", "Sarah Mohamed", "sarah@example.com", "201123456789", "Ali Hassan", "Gold", 1, "Director", "Creative Agency"],
            ["Sales Division", "Mohamed Ali", "mohamed@example.com", "201234567890", "Omar Said", "White", 0, "Manager", "Sales Corp"],
            ["Sales Division", "Fatma Nour", "fatma@example.com", "201345678901", "Layla Ahmed", "Gold", 1, "VP Sales", "Nile Group"],
        ]

        for row_idx, row_data in enumerate(sample_data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                template_ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            template_ws.column_dimensions[col].width = 22

        import tempfile
        template_path = os.path.join(tempfile.gettempdir(), 'admin_contacts_import_template.xlsx')
        wb.save(template_path)
        return template_path
